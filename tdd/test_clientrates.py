import unittest
import clientrates
import openpyxl
from openpyxl import Workbook, worksheet


class TestCR(unittest.TestCase):

    def test_get_rates(self):
        wb = self.create_test_wb()
        ratesDict = {'Desk':1, 'Gallery':5, 'Meeting':3, 'Theatre':20, 'Outdoor':20}
        self.assertEqual(ratesDict,clientrates.get_rates(wb,'Rates'))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        self.assertEqual(ratesDict,clientrates.get_rates(wb,'Rates'))   

    def test_facilitiesDictionary(self):
        wb = self.create_test_wb()
        fakeFacilities = {'Duck':'Pond', 'Horse':'Barn', 'Sheep':'Pen', 'Goose':'Pond', 'Goat':'Pen'}
        self.assertEqual(fakeFacilities,clientrates.facilitiesList(wb,'Facilities'))
        
        fakeFacilities = {'Horse':'Barn', 'Sheep':'Pen', 'Goose':'Pond', 'Goat':'Pen'}
        self.assertNotEqual(fakeFacilities,clientrates.facilitiesList(wb,'Facilities'))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        realFacilities = {'First Floor Hall':'Gallery',
                            'Second Floor Hall':'Gallery',
                            'Third Floor Hall':'Gallery',
                            'RGO Treehouse (South)':'Meeting',
                            'RGO Treehouse (North)':'Meeting',
                            'RGO Treehouls (Full)':'Meeting',
                            'Studio (North)':'Theatre',
                            'Studio (South)':'Theatre',
                            'ArtPark':'Outdoor',
                            'Desk 3rd floor 1':'Desk',
                            'Desk 3rd floor 2':'Desk',
                            'Desk 3rd floor 3':'Desk',
                            'Desk 3rd floor 4':'Desk',
                            'Desk 3rd floor 5':'Desk',
                            'Desk 3rd floor 6':'Desk',
                            'Desk 3rd floor 7':'Desk',
                            'Desk 3rd floor 8':'Desk',
                            'Desk 3rd floor 9':'Desk'}
        self.assertEqual(realFacilities,clientrates.facilitiesList(wb,'Facilities'))   

    def test_clientCreation(self):
        duckworth = clientrates.Client('duck','worth')
        self.assertEqual('duck',duckworth.first_name)
        self.assertEqual('worth',duckworth.last_name)
        self.assertEqual(0,duckworth.credits)
    
    def test_clientCredits(self):
        duckworth = clientrates.Client('duck','worth')
        self.assertEqual(0,duckworth.credits)
        duckworth.update_credits(5)
        self.assertEqual(5,duckworth.credits)
        duckworth.update_credits(5)
        self.assertEqual(10,duckworth.credits)
    
    def test_multipleClients(self):
        wb = self.create_test_wb()
        createdClients = clientrates.createClients(wb,'Clients')
        self.assertEqual('Sally', createdClients['Sally'].first_name)
        self.assertEqual(0, createdClients['Sally'].credits)
        self.assertEqual('Bob', createdClients['Bob'].first_name)
        self.assertEqual(0, createdClients['Bob'].credits)
        self.assertEqual('Slob', createdClients['Slob'].first_name)
        self.assertEqual(0, createdClients['Slob'].credits)

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        createdClients = clientrates.createClients(wb,'Clients')
        self.assertEqual('Jeramy', createdClients['Jeramy'].first_name)
        self.assertEqual('Beeler', createdClients['Jeramy'].last_name)
        self.assertEqual(0, createdClients['Jeramy'].credits)

    def test_monthlyClients(self):
        wb = self.create_test_wb()
        fakeClients = [['Sally','Room a'], ['Bob', 'Room b'], ['Slob', 'Room a']]
        self.assertEqual(fakeClients,clientrates.monthlyClients(wb,'2012-08'))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        self.assertIn(['Carri','Second Floor Hall'],clientrates.monthlyClients(wb,'2018-08'))
        self.assertNotIn(['Carri','Third Floor Hall'],clientrates.monthlyClients(wb,'2018-08'))
        self.assertIn(['Lorraine','RGO Treehouse (North)'],clientrates.monthlyClients(wb,'2018-08'))
        self.assertNotIn(['Geri','Third Floor Hall'],clientrates.monthlyClients(wb,'2018-08'))
        self.assertIn(['Geri','Desk 3rd floor 8'],clientrates.monthlyClients(wb,'2018-08'))

    def test_grabRates(self):
        wb = self.create_test_wb()
        rooms = clientrates.facilitiesList(wb,'Facilities')
        rates = clientrates.get_rates(wb,'AnimalRates')
        self.assertEqual(1,clientrates.grab_rate('Horse',rooms,rates))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        rooms = clientrates.facilitiesList(wb,'Facilities')
        rates = clientrates.get_rates(wb,'Rates')
        self.assertEqual(20,clientrates.grab_rate('ArtPark',rooms,rates))
    
    def test_clientCredits(self):
        # wb = self.create_test_wb()
        # self.assertEqual({'Sally':1},clientrates.calc_clientCredits(wb,'2012-09'))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        clientCredits = clientrates.calc_clientCredits(wb,'2018-08')
        self.assertEqual(54,clientCredits['Geri'])
        self.assertEqual(0,clientCredits['Jeramy'])

    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Rates")
        wb.create_sheet("AnimalRates")
        wb.create_sheet("Clients")
        wb.create_sheet("Facilities")
        wb.create_sheet("2012-08")
        wb.create_sheet("2012-09")

        ws = wb["Facilities"]
        ws.cell(row=1, column=1).value = 'Room Name'
        ws.cell(row=2, column=1).value = 'Duck'
        ws.cell(row=3, column=1).value = 'Horse'
        ws.cell(row=4, column=1).value = 'Sheep'
        ws.cell(row=5, column=1).value = 'Goose'
        ws.cell(row=6, column=1).value = 'Goat'

        ws.cell(row=1, column=2).value = 'Type'
        ws.cell(row=2, column=2).value = 'Pond'
        ws.cell(row=3, column=2).value = 'Barn'
        ws.cell(row=4, column=2).value = 'Pen'
        ws.cell(row=5, column=2).value = 'Pond'
        ws.cell(row=6, column=2).value = 'Pen'

        ws = wb["2012-08"]
        ws.cell(row=1, column=3).value = 'Room a'
        ws.cell(row=1, column=4).value = 'Room b'
        ws.cell(row=1, column=5).value = 'Room c'

        ws.cell(row=2, column=1).value = 'date'
        ws.cell(row=2, column=2).value = 'day'
        ws.cell(row=2, column=3).value = 'Sally'
        ws.cell(row=2, column=4).value = 'Bob'
        ws.cell(row=2, column=5).value = None

        ws.cell(row=5, column=3).value = 'Slob'

        ws = wb["2012-09"]
        ws.cell(row=1, column=3).value = 'Horse'
        ws.cell(row=1, column=4).value = 'Duck'
        ws.cell(row=1, column=5).value = 'Goose'

        ws.cell(row=2, column=1).value = 'date'
        ws.cell(row=2, column=2).value = 'day'
        ws.cell(row=2, column=3).value = 'Sally'
        ws.cell(row=2, column=4).value = 'Bob'
        ws.cell(row=2, column=5).value = None

        ws.cell(row=5, column=3).value = 'Slob'

        ws = wb['Rates']
        ws.cell(row=1, column=2).value = 'Type'
        ws.cell(row=2, column=2).value = 'Desk'
        ws.cell(row=3, column=2).value = 'Gallery'
        ws.cell(row=4, column=2).value = 'Meeting'
        ws.cell(row=5, column=2).value = 'Theatre'
        ws.cell(row=6, column=2).value = 'Outdoor'

        ws.cell(row=1, column=3).value = 'Credits'
        ws.cell(row=2, column=3).value = 1
        ws.cell(row=3, column=3).value = 5
        ws.cell(row=4, column=3).value = 3
        ws.cell(row=5, column=3).value = 20
        ws.cell(row=6, column=3).value = 20

        ws = wb['AnimalRates']
        ws.cell(row=1, column=2).value = 'Type'
        ws.cell(row=2, column=2).value = 'Barn'
        ws.cell(row=3, column=2).value = 'Pen'
        ws.cell(row=4, column=2).value = 'Pond'

        ws.cell(row=1, column=3).value = 'Credits'
        ws.cell(row=2, column=3).value = 1
        ws.cell(row=3, column=3).value = 5
        ws.cell(row=4, column=3).value = 3

        ws = wb['Clients']
        ws.cell(row=1, column=1).value = 'Name'
        ws.cell(row=2, column=1).value = 'Sally Ho'
        ws.cell(row=3, column=1).value = 'Bob Ng'
        ws.cell(row=4, column=1).value = 'Slob Bobby'

        # wb.save('test.xlsx')

        return wb