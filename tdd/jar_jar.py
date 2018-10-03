import openpyxl
import sys

from dateutil import parser

def get_year_month(bookingdate):
    return bookingdate[0:7]

def get_ymd(bookingdate):
	if bookingdate[8:]=='01':
		ymd = parser.parse(bookingdate)
	elif (int(bookingdate[8:]) < 10) and (int(bookingdate[8:]) >= 2):
		ymd = '=A' + bookingdate[9] + '+1'
	else:
		ymd = '=A' + bookingdate[8:] + '+1'
	return ymd

def getrooms(bookingdate):
	filename = 'cSpace_Bookingv1.xlsx'

	workbook = openpyxl.load_workbook(filename)
	ws = workbook[get_year_month(bookingdate)]
	
	# make a list out of Column A (the dates)
	dateList = []
	for cell in ws['A']:
		dateList.append(cell.value)
	
	# find the row that corresponds to the booking date
	selectRow = dateList.index(get_ymd(bookingdate)) + 1

	rowList = [] 
	roomList = []
	fullrooms = []

	# once a booking date is selected, create a list out of the desired row
	for cell in ws[selectRow]:
		rowList.append(cell.value)

	# create a list of all the rooms at cSpace
	for cell in ws[1]:
		roomList.append(cell.value)

	# create a list of full rooms at the desired date
	for i in range(len(rowList)):
		if rowList[i]:
			fullrooms.append(roomList[i])
	
	# remove all the full rooms from the room list
	for room in fullrooms:
		roomList.remove(room)

	return roomList

date=input("Please input the date using the format yyyy-mm-dd: ")
print(getrooms(date))



