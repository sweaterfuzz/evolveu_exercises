import unittest
import tdd2
import datetime
from openpyxl import Workbook, worksheet
import openpyxl

class TestStuff(unittest.TestCase):
    
    def test_add_five(self):
        self.assertEqual(8, tdd2.add_five(3))
    
    def test_my_max(self):
        self.assertEqual(5,tdd2.my_max([1,2,3,4,5]))
        self.assertEqual(-1,tdd2.my_max([-1,-2,-3,-4,-5]))
    
    def test_my_min(self):
        self.assertEqual(1,tdd2.my_min([1,2,3,4,5]))
        self.assertEqual(-5,tdd2.my_min([-1,-2,-3,-4,-5]))
    
    def test_has_string(self):
        self.assertEqual(["Mary had"],
						tdd2.has_string(
							["Mary had",
							 "a little lamb",
							 "little lamb",
							 "Whose fleece",
							],"Mary"))
        self.assertEqual(["a little lamb", "little lamb"],
						tdd2.has_string(
							["Mary had",
							 "a little lamb",
							 "little lamb",
							 "Whose fleece",
							],"amb"))

    def test_to_date(self):
        dt = tdd2.to_date("2010-08-02")
        self.assertIsInstance(dt, datetime.date)
        self.assertEqual(2010,dt.year)
        self.assertEqual(8,dt.month)
        self.assertEqual(2,dt.day)
    
    def test_date_diff(self):
        self.assertEqual(1, tdd2.date_diff("2018-09-02", "2018-09-01"))
        self.assertEqual(366, tdd2.date_diff("2019-09-02", "2018-09-01"))
        self.assertEqual(2080, tdd2.date_diff("2018-09-01", "2012-12-21"))
    
    def test_contains(self):
        self.assertTrue(tdd2.contains(['a','b','d'],"a"))
        self.assertFalse(tdd2.contains(['a','b','d'],"c"))

    def test_add_contents(self):
        self.assertEqual(6, tdd2.add_contents([1,2,3]))
        self.assertEqual(1, tdd2.add_contents([0,1,0]))
        self.assertEqual(6, tdd2.add_contents([-6,6,6]))
    
    def test_lookup(self):
        self.assertEqual('one mine', tdd2.lookup({1:'one', 2:'two', 3:'three'} , 1))

    def test_getClientList(self):
        wb = self.create_test_wb()
        self.assertEqual(['Sally','Bob','Slob'],tdd2.getClientList(wb))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        clientNames = ['Carri', 'Caren', 'Sierra', 'Delcie', 'Geri', 'Lavera',
                        'Soila', 'Valeri', 'Samira', 'Leone', 'Marylee', 'Annalisa',
                        'Dorathy', 'Denese', 'Larraine', 'Cassie', 'Sherrie', 
                        'Alicia', 'Jeramy', 'Clarissa']
        self.assertEqual(clientNames,tdd2.getClientList(wb))    

    def test_getMonthlyClients(self):
        wb = self.create_test_wb()
        self.assertEqual(['Sally','Bob','Fred','Slob'],tdd2.getMonthlyClients(wb,'2012-07'))    

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        self.assertIn('Carri',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertIn('Sherrie',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertIn('Lavera',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertIn('Geri',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertIn('Lorraine',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertIn('Clarissa',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Caren',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Sierra',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Delcie',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Soila',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Valeri',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Samira',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Leone',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Marylee',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Annalisa',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Dorathy',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Denese',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Larraine',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Cassie',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Alicia',tdd2.getMonthlyClients(wb,'2018-09'))
        self.assertNotIn('Jeramy',tdd2.getMonthlyClients(wb,'2018-09'))

    def test_getUnverifiedClients(self):
        list1 = ['Sally','Bob','Slob']
        list2 = ['Sally','Bob','Fred','Slob']
        self.assertEqual(['Fred'],tdd2.getUnverifiedClients(list1,list2))
        list1 = ['Sally','Bob','Slob']
        list2 = ['Sally','Bob','Fred','Slob','Dee']
        self.assertEqual(['Fred','Dee'],tdd2.getUnverifiedClients(list1,list2))
        self.assertEqual([],tdd2.getUnverifiedClients(list1,list1))

    def test_getFirstNames(self):
        self.assertEqual(['fred','sally','eve'],tdd2.getFirstNames(['fred ho','sally ng','eve eye']))

    def test_client_verify(self):
        wb = self.create_test_wb()
        self.assertFalse(tdd2.client_verify(wb,'2012-07'))
        self.assertTrue(tdd2.client_verify(wb,'2012-08'))

        filename = 'cSpace_Bookingv1.xlsx' 
        wb = openpyxl.load_workbook(filename)
        self.assertTrue(tdd2.client_verify(wb,'2018-07'))
        self.assertFalse(tdd2.client_verify(wb,'2018-08'))
        self.assertFalse(tdd2.client_verify(wb,'2018-09'))
        self.assertFalse(tdd2.client_verify(wb,'2018-10'))
        self.assertFalse(tdd2.client_verify(wb,'2018-11'))

    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Mysheet")
        wb.create_sheet("Clients")
        wb.create_sheet("2012-07")
        wb.create_sheet("2012-08")
        wb.create_sheet("2012-09")

        ws = wb["2012-07"]
        ws.cell(row=1, column=3).value = 'Room a'
        ws.cell(row=1, column=4).value = 'Room b'
        ws.cell(row=1, column=5).value = 'Room c'

        ws.cell(row=2, column=1).value = 'date'
        ws.cell(row=2, column=2).value = 'day'
        ws.cell(row=2, column=3).value = 'Sally'
        ws.cell(row=2, column=4).value = 'Bob'
        ws.cell(row=2, column=5).value = 'Fred'

        ws.cell(row=5, column=3).value = 'Slob'

        ws = wb["2012-08"]
        ws.cell(row=1, column=3).value = 'Room a'
        ws.cell(row=1, column=4).value = 'Room b'
        ws.cell(row=1, column=5).value = 'Room c'

        ws.cell(row=2, column=1).value = 'date'
        ws.cell(row=2, column=2).value = 'day'
        ws.cell(row=2, column=3).value = 'Sally'
        ws.cell(row=2, column=4).value = 'Bob'
        ws.cell(row=2, column=5).value = None

        ws.cell(row=5, column=3).value = 'Slob'


        ws = wb['Clients']
        ws.cell(row=1, column=1).value = 'Name'
        ws.cell(row=2, column=1).value = 'Sally Ho'
        ws.cell(row=3, column=1).value = 'Bob Ng'
        ws.cell(row=4, column=1).value = 'Slob Bobby'

        # wb.save('test.xlsx')

        return wb