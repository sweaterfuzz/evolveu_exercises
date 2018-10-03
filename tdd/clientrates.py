import openpyxl 
import pprint

# returns a dictionary of {'roomtype':rate} from the 'Rates' tab
def get_rates(wb,wsname):
    ws = wb[wsname]
    rateDict = {}

    for i in range(2,ws.max_row+1):
        roomtype = ws.cell(i,2).value
        credit = ws.cell(i,3).value
        rateDict[roomtype] = credit

    return rateDict

# returns a dictionary of {'roomname':'roomtype'} from the 'Facilities' tab
def facilitiesList(wb,wsname):
    ws = wb[wsname]
    facDict = {}

    for i in range(2,ws.max_row+1):
        roomname = ws.cell(i,1).value
        roomtype = ws.cell(i,2).value
        facDict[roomname] = roomtype

    return facDict

# grabs the rate (int) of a room (combining info from 'Facilities' tab and 'Rates' tab)
def grab_rate(room,allrooms,rates):
    typeOfRoom = allrooms[room]
    return rates[typeOfRoom]

# Client object with 3 attributes: first name, last name, room credits
class Client:

    def __init__(self,first_name,last_name):
        self.first_name = first_name
        self.last_name = last_name
        self.credits = 0

    def update_credits(self,c):
        self.credits = self.credits + c
        return self.credits

# creates a dictionary of Client objects {'first name': client Object} 
# from the 'Clients' tab
def createClients(wb,wsname):
    ws = wb[wsname]
    allClients = {}

    for i in range(2,ws.max_row+1):
        firstname, lastname = ws.cell(i,1).value.split()
        c = Client(firstname,lastname)
        allClients[firstname] = c 

    return allClients

# returns a list of [client,roomname] tuples for a given month
# (turns the monthly schedule into a list)
def monthlyClients(wb,wsname):
    ws = wb[wsname]
    clientList = []

    for i in range(2,ws.max_row+1):
        for j in range (3,ws.max_column):
            if ws.cell(i,j).value:
                clientName = ws.cell(i,j).value
                roomName = ws.cell(1,j).value
                clientList.append([clientName,roomName])
    
    return clientList

# calculate the number of credits a client has for a given month
def calc_clientCredits(wb,month):
    allClientVisits = monthlyClients(wb,month) # list[clientName,roomName]
    allClients = createClients(wb,'Clients') #dict{'first_name':Client object}
    facilities = facilitiesList(wb,'Facilities') #dict{'roomname':'roomtype'}
    ratesList = get_rates(wb,'Rates') #dict{'roomtype':int}

    for client in allClientVisits:
        name = client[0]
        room = client[1]
        rate = grab_rate(room,facilities,ratesList)
        allClients[name].update_credits(rate) 

    clientCredits = {}

    for k in allClients:
        clientCredits[k] = allClients[k].credits

    return clientCredits

# ----------------------------------------------------------------- #
filename = 'cSpace_Bookingv1.xlsx' 
wb = openpyxl.load_workbook(filename)
julyCredits = calc_clientCredits(wb,'2018-07')
augCredits = calc_clientCredits(wb,'2018-08')
septCredits = calc_clientCredits(wb,'2018-09')
octCredits = calc_clientCredits(wb,'2018-10')
novCredits = calc_clientCredits(wb,'2018-11')

pp = pprint.PrettyPrinter(indent=4)

print('July credit list:\n')
pp.pprint(julyCredits)
print('\n')

print('August credit list:\n')
pp.pprint(augCredits)
print('\n')

print('September credit list:\n')
pp.pprint(septCredits)
print('\n')

print('October credit list:\n')
pp.pprint(octCredits)
print('\n')

print('November credit list:\n')
pp.pprint(novCredits)
print('\n')